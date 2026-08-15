#!/usr/bin/env python3
"""生成 easy-agent 后端全部接口的 Excel 接口文档。

用法：
    .venv/bin/python scripts/generate_api_docs.py

输出：
    docs/API接口文档.xlsx

文档包含三个 Sheet：
    1. 接口总览   —— 每个接口一行：模块/方法/路径/名称/描述/鉴权/请求体/参数/响应
    2. 请求参数   —— 逐参数展开：Path/Query/Body 字段（含嵌套对象字段）
    3. 响应参数   —— 逐字段展开 200 响应结构
"""

from __future__ import annotations

import sys
import inspect
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from easy_agent.app import app  # noqa: E402
from easy_agent.middleware import get_current_username  # noqa: E402

OUTPUT = PROJECT_ROOT / "docs" / "API接口文档.xlsx"

# 不需要纳入文档的路由（框架/页面路由）
EXCLUDE_PATHS = {
    "/docs",
    "/docs/oauth2-redirect",
    "/openapi.json",
    "/swagger-ui-assets",
    "/terminal",
    "/",
    "/{full_path:path}",
    "/{full_path}",
}

MODULE_LABELS = [
    ("/api/auth", "认证"),
    ("/api/chat", "聊天"),
    ("/api/sessions", "会话"),
    ("/api/files", "文件"),
    ("/api/settings", "设置"),
    ("/api/skill-center", "技能中心"),
    ("/api/scheduled-tasks", "定时任务"),
    ("/api/bloom", "彭博分析"),
    ("/api/forex", "外汇"),
    ("/api/prompts", "提示词"),
    ("/api/completion", "对话补全"),
    ("/api/health", "健康检查"),
    ("/api/config", "配置"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def module_of(path: str) -> str:
    for prefix, label in MODULE_LABELS:
        if path.startswith(prefix):
            return label
    return "其他"


def auth_required(route: Any) -> bool:
    for dep in route.dependant.dependencies:
        fn = getattr(dep.call, "__func__", dep.call)
        if getattr(fn, "__name__", "") == get_current_username.__name__:
            return True
    return False


def type_str(schema: dict | None, openapi: dict) -> str:
    if not isinstance(schema, dict):
        return "any"
    if "$ref" in schema:
        return schema["$ref"].split("/")[-1]
    t = schema.get("type")
    fmt = schema.get("format")
    if schema.get("contentMediaType"):
        return f"file(binary: {schema['contentMediaType']})"
    if fmt == "binary":
        return "file(binary)"
    if t == "array":
        return f"array<{type_str(schema.get('items'), openapi)}>"
    if t == "string" and schema.get("enum"):
        return "string(枚举: " + "|".join(str(e) for e in schema["enum"]) + ")"
    if t is None and ("anyOf" in schema or "oneOf" in schema):
        variants = schema.get("anyOf") or schema.get("oneOf")
        return " | ".join(type_str(v, openapi) for v in variants)
    base = t or "object"
    if fmt:
        base = f"{base}({fmt})"
    return base


def _resolve(schema: dict | None, openapi: dict) -> dict:
    seen = 0
    while schema and "$ref" in schema and seen < 10:
        name = schema["$ref"].split("/")[-1]
        schema = openapi.get("components", {}).get("schemas", {}).get(name)
        seen += 1
    return schema or {}


def flatten_schema(
    schema: dict | None,
    openapi: dict,
    prefix: str = "",
    depth: int = 0,
    seen: set | None = None,
) -> list[dict]:
    """把 JSON Schema 展平为字段行（含嵌套对象，最多 8 层）。"""
    if depth > 8:
        return []
    seen = seen or set()
    schema = _resolve(schema, openapi)
    if not schema:
        return []

    if "allOf" in schema:
        rows: list[dict] = []
        for sub in schema["allOf"]:
            rows.extend(flatten_schema(sub, openapi, prefix, depth, seen))
        return rows

    if schema.get("type") == "array":
        items = _resolve(schema.get("items"), openapi)
        if items.get("type") == "object" and items.get("properties"):
            return flatten_schema(items, openapi, prefix, depth, seen)
        return []

    if schema.get("type") != "object" or not schema.get("properties"):
        return []

    required = set(schema.get("required") or [])
    rows = []
    for name, prop in schema["properties"].items():
        prop = _resolve(prop, openapi)
        field = f"{prefix}{name}"
        rows.append({
            "field": field,
            "type": type_str(prop, openapi),
            "required": name in required,
            "default": prop.get("default", ""),
            "desc": prop.get("description", "") or "",
        })
        child = _resolve(prop, openapi)
        if child.get("type") == "object" and child.get("properties"):
            rows.extend(
                flatten_schema(child, openapi, f"{field}.", depth + 1, seen)
            )
        elif child.get("type") == "array":
            items = _resolve(child.get("items"), openapi)
            if items.get("type") == "object" and items.get("properties"):
                rows.extend(
                    flatten_schema(items, openapi, f"{field}[].", depth + 1, seen)
                )
    return rows


def request_body_rows(operation: dict, openapi: dict) -> tuple[str, list[dict]]:
    """返回 (请求体概述, 请求字段行)。"""
    rb = operation.get("requestBody") or {}
    content = rb.get("content") or {}
    if not content:
        return "", []
    media = next(iter(content.values()))
    schema = media.get("schema")
    if not schema:
        return ", ".join(content.keys()), []
    media_types = ", ".join(content.keys())
    rows = flatten_schema(schema, openapi)
    ref = schema.get("$ref", "")
    label = ref.split("/")[-1] if ref else type_str(schema, openapi)
    return f"{label} ({media_types})", rows


def param_rows(operation: dict, openapi: dict) -> list[dict]:
    rows = []
    for p in operation.get("parameters") or []:
        schema = p.get("schema") or {}
        rows.append({
            "loc": (p.get("in") or "").upper(),
            "field": p.get("name", ""),
            "type": type_str(schema, openapi),
            "required": bool(p.get("required")),
            "default": schema.get("default", ""),
            "desc": p.get("description", "") or "",
        })
    return rows


def response_rows(operation: dict, openapi: dict, is_sse: bool = False) -> tuple[str, list[dict]]:
    if is_sse:
        return "SSE 流式事件 (text/event-stream)", []
    responses = operation.get("responses") or {}
    resp = responses.get("200") or responses.get("201") or {}
    content = resp.get("content") or {}
    if not content:
        return "", []
    if "text/event-stream" in content:
        return "SSE 流式事件", []
    media = next(iter(content.values()))
    schema = media.get("schema")
    if not schema:
        return ", ".join(content.keys()), []
    rows = flatten_schema(schema, openapi)
    ref = schema.get("$ref", "")
    label = ref.split("/")[-1] if ref else type_str(schema, openapi)
    return f"{label} ({', '.join(content.keys())})", rows


def collect_endpoints() -> list[dict]:
    openapi = app.openapi()
    endpoints = []
    seq = 0
    for path, methods in openapi.get("paths", {}).items():
        if path in EXCLUDE_PATHS:
            continue
        route = next(
            (r for r in app.routes if getattr(r, "path", None) == path),
            None,
        )
        is_sse = False
        if route is not None:
            try:
                src = inspect.getsource(route.endpoint)
                is_sse = "StreamingResponse(" in src
            except Exception:
                is_sse = route.response_class is StreamingResponse
        for method, op in methods.items():
            if method not in ("get", "post", "put", "patch", "delete"):
                continue
            seq += 1
            body_label, body_rows = request_body_rows(op, openapi)
            params = param_rows(op, openapi)
            resp_label, resp_rows = response_rows(op, openapi, is_sse=is_sse)
            need_auth = auth_required(route) if route is not None else False
            endpoints.append({
                "seq": seq,
                "module": module_of(path),
                "method": method.upper(),
                "path": path,
                "name": op.get("summary", "") or op.get("operationId", ""),
                "desc": (op.get("description") or "").strip(),
                "auth": "是" if need_auth else "否",
                "body_label": body_label,
                "body_rows": body_rows,
                "params": params,
                "resp_label": resp_label,
                "resp_rows": resp_rows,
            })
    endpoints.sort(key=lambda e: (e["module"], e["path"], e["method"]))
    for i, e in enumerate(endpoints, 1):
        e["no"] = i
    return endpoints


def style_sheet(ws, widths: list[int], n_rows: int, n_cols: int) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1, max_col=n_cols):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"


def main() -> int:
    endpoints = collect_endpoints()
    wb = Workbook()

    # ── Sheet 1: 接口总览 ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "接口总览"
    headers1 = ["序号", "模块", "方法", "路径", "接口名称", "功能描述", "鉴权", "请求体", "请求参数(Query/Path)", "响应说明"]
    ws1.append(headers1)
    for e in endpoints:
        param_summary = "; ".join(
            f"{p['field']}({'必填' if p['required'] else '选填'})"
            for p in e["params"]
        )
        ws1.append([
            e["no"],
            e["module"],
            e["method"],
            e["path"],
            e["name"],
            e["desc"],
            e["auth"],
            e["body_label"],
            param_summary,
            e["resp_label"],
        ])
    style_sheet(ws1, [6, 10, 7, 38, 22, 45, 6, 24, 30, 26], len(endpoints), len(headers1))

    # ── Sheet 2: 请求参数 ──────────────────────────────────────────────
    ws2 = wb.create_sheet("请求参数")
    headers2 = ["序号", "方法", "路径", "参数位置", "参数名", "类型", "必填", "默认值", "说明"]
    ws2.append(headers2)
    row_idx = 0
    for e in endpoints:
        for p in e["params"]:
            row_idx += 1
            ws2.append([e["no"], e["method"], e["path"], p["loc"], p["field"], p["type"],
                        "是" if p["required"] else "否", p["default"], p["desc"]])
        for b in e["body_rows"]:
            row_idx += 1
            ws2.append([e["no"], e["method"], e["path"], "BODY", b["field"], b["type"],
                        "是" if b["required"] else "否", b["default"], b["desc"]])
    style_sheet(ws2, [6, 7, 40, 10, 32, 22, 6, 14, 45], row_idx, len(headers2))

    # ── Sheet 3: 响应参数 ──────────────────────────────────────────────
    ws3 = wb.create_sheet("响应参数")
    headers3 = ["序号", "方法", "路径", "响应字段", "类型", "必填", "说明"]
    ws3.append(headers3)
    row_idx = 0
    for e in endpoints:
        for r in e["resp_rows"]:
            row_idx += 1
            ws3.append([e["no"], e["method"], e["path"], r["field"], r["type"],
                        "是" if r["required"] else "否", r["desc"]])
    style_sheet(ws3, [6, 7, 40, 36, 22, 6, 45], row_idx, len(headers3))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"已生成 {OUTPUT}（接口 {len(endpoints)} 个）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
